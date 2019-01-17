from openpyxl import load_workbook
from material import Material
from piec import Piec
wb = load_workbook('template.xlsx')
print(wb.sheetnames)


def set_funkcja_celu(materials):
    counts=0.0
    costs=0.0
    for m in materials:
        costs *= m.cost
        counts *= m.count
    return costs+counts


materials = []
sheet = wb.get_sheet_by_name("Materialy")
for i in range(2,10):
    if sheet.cell(column=1, row=i).value:
        materials.append(Material(sheet.cell(column=1, row=i).value, sheet.cell(column=2, row=i).value))

for m in materials:
    print(m)

materials_count = len(materials)
print("Komorka celu: "+ str(set_funkcja_celu(materials)))


sheet = wb.get_sheet_by_name("Czas produkcji")

piece = []
for i in range(1,10):
    time_per_material=[]
    if sheet.cell(column=i, row=2).value:
        piec=Piec(sheet.cell(column=i, row=2).value)
        for j in range(3,13):
            if sheet.cell(column=i, row=j).value:
                time_per_material.append(sheet.cell(column=i, row=j).value)
        piec.time_per_material=time_per_material
        piece.append(piec)


piec_count = len(piece)

sheet = wb.get_sheet_by_name("Limit czasu")
for i in range(2, 2+piec_count):
    if sheet.cell(column=1, row=i).value:
        piece[i-2].set_time_limit(sheet.cell(column=2, row=i).value)


for p in piece:
    print(p)